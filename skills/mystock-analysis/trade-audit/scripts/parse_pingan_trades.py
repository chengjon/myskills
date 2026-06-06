#!/usr/bin/env python3
"""
平安证券交易对账单解析器
解析 pingan*.xlsx，普通/两融分别保存到MySQL，汇总到Obsidian
"""

import openpyxl
import os
import re
import sys
import json
from datetime import datetime
from decimal import Decimal

# ============================================================
# 配置
# ============================================================
DATA_DIR = "/mnt/d/MyCode3/1000w/data/pinan/"
OB_DIR = "/mnt/c/Users/John Cheng/Documents/Obsidian Vault/mystocks/历史交易/"

MYSQL_CONFIG = {
    "host": os.environ.get("MYSQL_HOST", ""),
    "port": 3306,
    "user": "root",
    "password": "",
    "database": "hermes",
    "charset": "utf8mb4",
}

# 普通交易列定义 (L8, 19列含末尾None)
NORMAL_COLS = [
    "posting_date",      # 记账日期
    "abstract",          # 业务摘要
    "account",           # 资产账户
    "shareholder_acct",  # 股东账号
    "stock_code",        # 证券代码
    "stock_name",        # 证券名称
    "shares",            # 成交股数
    "price",             # 成交价格
    "amount",            # 成交金额
    "interest",          # 利息金额
    "movement",          # 发生金额
    "balance",           # 资金余额
    "commission",        # 手续费
    "stamp_tax",         # 印花税
    "transfer_fee",      # 过户费等
    "trade_levy",        # 财汇局交易征费
    "currency",          # 币种
    "hk_fx_rate",        # 港股汇率
]

# 两融交易列定义 (17列+末尾可能有None)
MARGIN_COLS = [
    "posting_date",      # 记账日期
    "abstract",          # 业务摘要
    "account",           # 资产账户
    "shareholder_acct",  # 股东账号
    "stock_code",        # 证券代码
    "stock_name",        # 证券名称
    "shares",            # 成交股数
    "price",             # 成交价格
    "amount",            # 成交金额
    "interest",          # 利息金额
    "movement",          # 发生金额
    "balance",           # 资金余额
    "commission",        # 手续费
    "stamp_tax",         # 印花税
    "transfer_fee",      # 过户费
    "other_fee",         # 其他费
    "currency",          # 币种
]


def parse_xlsx(filepath):
    """解析单个xlsx文件，返回 {normal: [...], margin: [...], meta: {...}}"""
    fname = os.path.basename(filepath)
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # 1. 解析文件头(L1-4)提取元信息
    meta = {"filename": fname, "account_holder": "", "fund_account": "", "date_range": ""}
    rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
    for row in rows:
        for cell in (row or []):
            text = str(cell or "")
            if "打印日期范围" in text:
                m = re.search(r"(\d{4}-\d{2}-\d{2})~(\d{4}-\d{2}-\d{2})", text)
                if m:
                    meta["date_range"] = f"{m.group(1)}~{m.group(2)}"
            if "客户姓名" in text:
                m = re.search(r"Name[：:]\s*(.+)", text)
                if m:
                    meta["account_holder"] = m.group(1).strip()
            if "主资金账户" in text:
                m = re.search(r"Main Fund Account[：:]\s*(\d+)", text)
                if m:
                    meta["fund_account"] = m.group(1)

    # 2. 扫描全表，定位section边界
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    wb.close()

    normal_header_row = None   # 普通交易列头行号
    margin_header_row = None   # 两融交易列头行号
    summary_row = None         # "二、股份交易汇总"行号

    for i, row in enumerate(all_rows):
        v = str(row[0] or "")
        if "1、普通证券交易" in v and normal_header_row is None:
            normal_header_row = i  # section头，数据从i+2开始(跳过列头行+可能的空行)
        if "2、融资融券交易" in v and margin_header_row is None:
            margin_header_row = i
        if v.startswith("二、") and summary_row is None:
            summary_row = i

    # 3. 解析普通交易数据
    normal = []
    if normal_header_row is not None:
        # 列头在section头下一行，数据再下一行
        data_start = normal_header_row + 2  # 跳过section头+列头
        data_end = margin_header_row if margin_header_row else summary_row if summary_row else len(all_rows)
        for i in range(data_start, data_end):
            row = all_rows[i]
            if not row or not row[0]:
                continue
            record = _parse_normal_row(row, meta)
            if record:
                normal.append(record)

    # 4. 解析两融交易数据
    margin = []
    if margin_header_row is not None:
        data_start = margin_header_row + 2  # 跳过section头+列头
        data_end = summary_row if summary_row else len(all_rows)
        for i in range(data_start, data_end):
            row = all_rows[i]
            if not row or not row[0]:
                continue
            record = _parse_margin_row(row, meta)
            if record:
                margin.append(record)

    return {"normal": normal, "margin": margin, "meta": meta}


def _parse_normal_row(row, meta):
    """解析普通交易行"""
    if len(row) < 17:
        return None
    # 跳过空行、小计行、合计行
    date_str = str(row[0] or "").strip()
    if not date_str or not re.match(r"^\d{8}$", date_str):
        return None
    abstract = str(row[1] or "").strip()
    if abstract in ("小计", "合计"):
        return None

    record = {}
    for j, col in enumerate(NORMAL_COLS):
        if j < len(row):
            record[col] = row[j]
        else:
            record[col] = None

    # 清理: 去掉末尾None列
    record["source_file"] = meta["filename"]
    record["date_range"] = meta["date_range"]
    record["account_type"] = "normal"

    # 标准化日期 YYYYMMDD -> YYYY-MM-DD
    record["posting_date"] = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"

    # 数值字段转Decimal
    for fld in ["shares", "price", "amount", "interest", "movement", "balance",
                "commission", "stamp_tax", "transfer_fee", "trade_levy"]:
        record[fld] = _to_decimal(record.get(fld))

    return record


def _parse_margin_row(row, meta):
    """解析两融交易行"""
    if len(row) < 15:
        return None
    date_str = str(row[0] or "").strip()
    # 两融日期格式可能是 YYYYMMDD 或 YYYY-MM-DD
    if not date_str:
        return None
    if re.match(r"^\d{4}-\d{2}-\d{2}$", date_str):
        pass  # already good
    elif re.match(r"^\d{8}$", date_str):
        date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
    else:
        return None

    abstract = str(row[1] or "").strip()
    if abstract in ("小计", "合计"):
        return None

    record = {}
    for j, col in enumerate(MARGIN_COLS):
        if j < len(row):
            record[col] = row[j]
        else:
            record[col] = None

    record["source_file"] = meta["filename"]
    record["date_range"] = meta["date_range"]
    record["account_type"] = "margin"
    record["posting_date"] = date_str
    record["hk_fx_rate"] = None  # 两融没有港股汇率列
    record["trade_levy"] = None  # 两融没有财汇局征费列

    for fld in ["shares", "price", "amount", "interest", "movement", "balance",
                "commission", "stamp_tax", "transfer_fee", "other_fee"]:
        record[fld] = _to_decimal(record.get(fld))

    return record


def _to_decimal(val):
    """安全转Decimal"""
    if val is None:
        return Decimal("0")
    try:
        s = str(val).strip().replace(",", "")
        if not s or s == "None":
            return Decimal("0")
        return Decimal(s)
    except Exception:
        return Decimal("0")


# ============================================================
# MySQL建表与写入
# ============================================================
def ensure_mysql_tables():
    """创建MySQL表(如不存在)"""
    import pymysql
    conn = pymysql.connect(**MYSQL_CONFIG)
    cur = conn.cursor()

    # 普通交易明细表
    cur.execute("""
    CREATE TABLE IF NOT EXISTS pingan_normal_trade (
        id BIGINT AUTO_INCREMENT PRIMARY KEY,
        posting_date DATE NOT NULL,
        abstract VARCHAR(50),
        account VARCHAR(30),
        shareholder_acct VARCHAR(20),
        stock_code VARCHAR(10),
        stock_name VARCHAR(30),
        shares DECIMAL(18,2),
        price DECIMAL(12,3),
        amount DECIMAL(18,2),
        interest DECIMAL(18,2),
        movement DECIMAL(18,2),
        balance DECIMAL(18,2),
        commission DECIMAL(12,2),
        stamp_tax DECIMAL(12,2),
        transfer_fee DECIMAL(12,2),
        trade_levy DECIMAL(12,2),
        currency VARCHAR(10),
        hk_fx_rate DECIMAL(10,5),
        source_file VARCHAR(80),
        date_range VARCHAR(30),
        account_type VARCHAR(10) DEFAULT 'normal',
        UNIQUE KEY uk_normal_trade (posting_date, abstract, account, stock_code, shares, price, amount),
        KEY idx_normal_date (posting_date),
        KEY idx_normal_code (stock_code),
        KEY idx_normal_abstract (abstract)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)

    # 两融交易明细表
    cur.execute("""
    CREATE TABLE IF NOT EXISTS pingan_margin_trade (
        id BIGINT AUTO_INCREMENT PRIMARY KEY,
        posting_date DATE NOT NULL,
        abstract VARCHAR(50),
        account VARCHAR(30),
        shareholder_acct VARCHAR(20),
        stock_code VARCHAR(10),
        stock_name VARCHAR(30),
        shares DECIMAL(18,2),
        price DECIMAL(12,3),
        amount DECIMAL(18,2),
        interest DECIMAL(18,2),
        movement DECIMAL(18,2),
        balance DECIMAL(18,2),
        commission DECIMAL(12,2),
        stamp_tax DECIMAL(12,2),
        transfer_fee DECIMAL(12,2),
        other_fee DECIMAL(12,2),
        currency VARCHAR(10),
        source_file VARCHAR(80),
        date_range VARCHAR(30),
        account_type VARCHAR(10) DEFAULT 'margin',
        UNIQUE KEY uk_margin_trade (posting_date, abstract, account, stock_code, shares, price, amount),
        KEY idx_margin_date (posting_date),
        KEY idx_margin_code (stock_code),
        KEY idx_margin_abstract (abstract)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)

    conn.commit()
    cur.close()
    conn.close()
    print("  MySQL表已就绪")


def write_to_mysql(records, table_name):
    """批量写入MySQL，先清空再插入(保证完整)"""
    if not records:
        print(f"  {table_name}: 0条，跳过")
        return 0

    import pymysql
    conn = pymysql.connect(**MYSQL_CONFIG)
    cur = conn.cursor()

    # 根据表名确定列
    if table_name == "pingan_normal_trade":
        cols = NORMAL_COLS + ["source_file", "date_range", "account_type"]
    else:
        cols = MARGIN_COLS + ["source_file", "date_range", "account_type"]

    # 先清空旧数据，全量重写
    cur.execute(f"TRUNCATE TABLE {table_name}")

    placeholders = ", ".join(["%s"] * len(cols))
    col_names = ", ".join(f"`{c}`" for c in cols)
    sql = f"INSERT INTO {table_name} ({col_names}) VALUES ({placeholders})"

    inserted = 0
    failed = 0
    for rec in records:
        values = []
        for c in cols:
            v = rec.get(c)
            if isinstance(v, Decimal):
                v = float(v)
            values.append(v)
        try:
            cur.execute(sql, values)
            if cur.rowcount > 0:
                inserted += 1
        except Exception as e:
            failed += 1
            if failed <= 3:
                print(f"  写入失败: {rec['posting_date']} {rec.get('abstract','')} {rec.get('stock_code','')} - {e}")

    if failed > 3:
        print(f"  ... 共{failed}条失败")

    conn.commit()
    cur.close()
    conn.close()
    print(f"  {table_name}: 写入{inserted}条, 失败{failed}条")
    return inserted


# ============================================================
# Obsidian汇总报告
# ============================================================
def generate_ob_summary(normal_all, margin_all, meta_list):
    """生成Obsidian汇总markdown"""
    os.makedirs(OB_DIR, exist_ok=True)

    # ---- 1. 普通账户汇总 ----
    normal_trades = [r for r in normal_all if r.get("stock_code") and r["abstract"] in 
                     ("买入", "卖出", "证券买入清算", "证券卖出清算")]
    margin_trades = [r for r in margin_all if r.get("stock_code") and r["abstract"] in 
                     ("买入", "卖出", "证券买入清算", "证券卖出清算", 
                      "融资买入", "融券卖出", "买券还券", "卖券还款", 
                      "担保品买入", "担保品卖出",
                      "证券买入清算(融资买入)", "证券卖出清算(卖券还款)")]

    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    # 按年汇总普通账户
    _write_account_md("平安普通", normal_all, normal_trades, now, OB_DIR)
    # 按年汇总两融账户
    _write_account_md("平安两融", margin_all, margin_trades, now, OB_DIR)

    # ---- 综合概览 ----
    BUY_ABSTRACTS = {"买入", "证券买入清算", "融资买入", "担保品买入", "证券买入清算(融资买入)"}
    SELL_ABSTRACTS = {"卖出", "证券卖出清算", "融券卖出", "卖券还款", "担保品卖出", "证券卖出清算(卖券还款)"}
    n_buy = sum(1 for r in normal_trades if r["abstract"] in BUY_ABSTRACTS or (r["abstract"] not in SELL_ABSTRACTS and r["shares"] > 0))
    n_sell = sum(1 for r in normal_trades if r["abstract"] in SELL_ABSTRACTS or (r["abstract"] not in BUY_ABSTRACTS and r["shares"] < 0))
    m_buy = sum(1 for r in margin_trades if r["abstract"] in BUY_ABSTRACTS or (r["abstract"] not in SELL_ABSTRACTS and r["shares"] > 0))
    m_sell = sum(1 for r in margin_trades if r["abstract"] in SELL_ABSTRACTS or (r["abstract"] not in BUY_ABSTRACTS and r["shares"] < 0))

    n_commission = sum(abs(float(r["commission"])) for r in normal_trades)
    n_stamp = sum(abs(float(r["stamp_tax"])) for r in normal_trades)
    m_commission = sum(abs(float(r["commission"])) for r in margin_trades)
    m_stamp = sum(abs(float(r["stamp_tax"])) for r in margin_trades)

    md = f"""---
title: 平安证券交易汇总
date: {now}
accounts: 平安普通+平安两融
---

# 平安证券交易汇总

> 生成时间: {now}

## 账户概览

| 项目 | 普通账户 | 两融账户 | 合计 |
|------|---------|---------|------|
| 总明细条数 | {len(normal_all)} | {len(margin_all)} | {len(normal_all)+len(margin_all)} |
| 买入笔数 | {n_buy} | {m_buy} | {n_buy+m_buy} |
| 卖出笔数 | {n_sell} | {m_sell} | {n_sell+m_sell} |
| 手续费总额 | {n_commission:,.2f} | {m_commission:,.2f} | {n_commission+m_commission:,.2f} |
| 印花税总额 | {n_stamp:,.2f} | {m_stamp:,.2f} | {n_stamp+m_stamp:,.2f} |

## 数据来源

| 文件 | 日期范围 |
|------|---------|
"""
    for m in meta_list:
        md += f"| {m['filename']} | {m['date_range']} |\n"

    md += """
## 明细文件

- [[平安普通-交易明细]]
- [[平安两融-交易明细]]
"""
    path = os.path.join(OB_DIR, "平安证券交易汇总.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write(md)
    print(f"  汇总: {path}")


def _write_account_md(label, all_records, trade_records, now, ob_dir):
    """生成单个账户的Obsidian明细文件"""
    # 按股票汇总
    stock_summary = {}
    # 买入类摘要
    BUY_ABSTRACTS = {"买入", "证券买入清算", "融资买入", "担保品买入", "证券买入清算(融资买入)"}
    SELL_ABSTRACTS = {"卖出", "证券卖出清算", "融券卖出", "卖券还款", "担保品卖出", "证券卖出清算(卖券还款)"}
    
    for r in trade_records:
        code = r.get("stock_code") or ""
        if not code:
            continue
        name = r.get("stock_name") or code
        key = f"{code} {name}"
        if key not in stock_summary:
            stock_summary[key] = {"code": code, "name": name, "buy_count": 0, "sell_count": 0,
                                   "buy_shares": Decimal(0), "sell_shares": Decimal(0),
                                   "buy_amount": Decimal(0), "sell_amount": Decimal(0),
                                   "total_commission": Decimal(0), "total_stamp": Decimal(0)}
        s = stock_summary[key]
        abstract = r.get("abstract", "")
        is_buy = abstract in BUY_ABSTRACTS or (abstract not in SELL_ABSTRACTS and r["shares"] > 0)
        is_sell = abstract in SELL_ABSTRACTS or (abstract not in BUY_ABSTRACTS and r["shares"] < 0)
        
        if is_buy:
            s["buy_count"] += 1
            s["buy_shares"] += abs(r["shares"])
            s["buy_amount"] += abs(r["amount"])
        if is_sell:
            s["sell_count"] += 1
            s["sell_shares"] += abs(r["shares"])
            s["sell_amount"] += abs(r["amount"])
        s["total_commission"] += abs(r["commission"])
        s["total_stamp"] += abs(r["stamp_tax"])

    # 按年份分组
    year_groups = {}
    for r in all_records:
        yr = r["posting_date"][:4]
        year_groups.setdefault(yr, []).append(r)

    filename = f"{label}-交易明细.md"
    md = f"""---
title: {label}交易明细
date: {now}
account: {label}
---

# {label}交易明细

> 生成时间: {now}

## 按年统计

| 年份 | 明细条数 | 买入笔数 | 卖出笔数 |
|------|---------|---------|---------|
"""
    for yr in sorted(year_groups.keys()):
        recs = year_groups[yr]
        buys = sum(1 for r in recs if r["abstract"] in ("买入", "证券买入清算", "融资买入", "担保品买入") and r.get("stock_code"))
        sells = sum(1 for r in recs if r["abstract"] in ("卖出", "证券卖出清算", "融券卖出", "担保品卖出") and r.get("stock_code"))
        md += f"| {yr} | {len(recs)} | {buys} | {sells} |\n"

    if stock_summary:
        md += f"""
## 按股票汇总 ({len(stock_summary)}只)

| 股票 | 买入笔数 | 买入股数 | 买入金额 | 卖出笔数 | 卖出股数 | 卖出金额 | 手续费 | 印花税 |
|------|---------|---------|---------|---------|---------|---------|--------|--------|
"""
        for key in sorted(stock_summary.keys()):
            s = stock_summary[key]
            md += f"| {key} | {s['buy_count']} | {s['buy_shares']:,.0f} | {s['buy_amount']:,.2f} | {s['sell_count']} | {s['sell_shares']:,.0f} | {s['sell_amount']:,.2f} | {s['total_commission']:,.2f} | {s['total_stamp']:,.2f} |\n"

    md += f"""
## 年度明细

"""
    for yr in sorted(year_groups.keys()):
        recs = year_groups[yr]
        md += f"### {yr}年 ({len(recs)}条)\n\n"
        md += "| 日期 | 摘要 | 代码 | 名称 | 股数 | 价格 | 金额 | 手续费 | 印花税 | 发生金额 | 余额 |\n"
        md += "|------|------|------|------|------|------|------|--------|--------|---------|------|\n"
        for r in recs:
            code = r.get("stock_code") or ""
            name = r.get("stock_name") or ""
            md += f"| {r['posting_date']} | {r['abstract']} | {code} | {name} | {r['shares']:,.0f} | {r['price']:.3f} | {r['amount']:,.2f} | {r['commission']:,.2f} | {r['stamp_tax']:,.2f} | {r['movement']:,.2f} | {r['balance']:,.2f} |\n"

    path = os.path.join(ob_dir, filename)
    with open(path, "w", encoding="utf-8") as f:
        f.write(md)
    print(f"  明细: {path}")


# ============================================================
# 主流程
# ============================================================
def main():
    print("=" * 60)
    print("平安证券对账单解析器")
    print("=" * 60)

    # 1. 扫描文件
    files = sorted(f for f in os.listdir(DATA_DIR) if f.startswith("pingan") and f.endswith(".xlsx"))
    print(f"\n找到{len(files)}个文件:")
    for f in files:
        print(f"  {f}")

    # 2. 解析所有文件
    normal_all = []
    margin_all = []
    meta_list = []
    for fname in files:
        print(f"\n解析: {fname}...")
        result = parse_xlsx(os.path.join(DATA_DIR, fname))
        normal_all.extend(result["normal"])
        margin_all.extend(result["margin"])
        meta_list.append(result["meta"])
        print(f"  普通: {len(result['normal'])}条, 两融: {len(result['margin'])}条")

    print(f"\n合计: 普通{len(normal_all)}条, 两融{len(margin_all)}条")

    # 3. 建MySQL表
    print("\n[1/3] 建MySQL表...")
    ensure_mysql_tables()

    # 4. 写MySQL
    print("\n[2/3] 写MySQL...")
    n_count = write_to_mysql(normal_all, "pingan_normal_trade")
    m_count = write_to_mysql(margin_all, "pingan_margin_trade")

    # 5. 生成Obsidian汇总
    print("\n[3/3] 生成Obsidian汇总...")
    generate_ob_summary(normal_all, margin_all, meta_list)

    print("\n" + "=" * 60)
    print(f"完成! 普通账户{n_count}条 + 两融账户{m_count}条已入库")
    print(f"Obsidian汇总: {OB_DIR}")
    print("=" * 60)


if __name__ == "__main__":
    main()
